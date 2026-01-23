import streamlit as st
import pandas as pd
import numpy as np
import re
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Transaction Processing Model", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
.block-container {padding-top: 1rem; padding-bottom: 1rem;}
.main-header-container {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    padding: 1.5rem; border-radius: 10px; border: 3px solid #5a67d8;
    box-shadow: 0 4px 6px rgba(0,0,0,0.1); margin-bottom: 0.5rem; width: 100%;
}
.main-header {text-align: center; font-size: 2.5rem; font-weight: 700; color: white; margin: 0; padding: 0;}
.subtitle {text-align: center; font-size: 1.1rem; color: #555; margin-top: 1rem; margin-bottom: 2rem;}
.timer-box {text-align: right; padding: 12px 18px; background-color: #f0f2f6;
    border-radius: 8px; border: 2px solid #e1e4e8; margin-top: 5px;}
.timer-label {font-size: 0.85rem; color: #666; margin-bottom: 3px;}
.timer-value {font-size: 1.5rem; font-weight: 700; color: #1f77b4;}
.stMetric {background-color: #f8f9fa; border: 2px solid #e1e4e8; padding: 15px; border-radius: 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);}
</style>
""", unsafe_allow_html=True)

def normalize_col(c):
    return re.sub(r'[^a-z0-9]', '', str(c).lower())

def find_col(df, possible):
    col_map = {normalize_col(c): c for c in df.columns}
    for p in possible:
        if normalize_col(p) in col_map:
            return col_map[normalize_col(p)]
    raise Exception(f"Missing column. Tried {possible}")

def strip_time_from_dates(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.date
    return df

# Initialize timer placeholder
if 'timer_placeholder' not in st.session_state:
    st.session_state.timer_placeholder = None

header_col, timer_col = st.columns([5, 1])
with header_col:
    st.markdown('<div class="main-header-container"><h1 class="main-header">📊 Transaction Processing Model</h1></div>', unsafe_allow_html=True)
with timer_col:
    if 'start_time' not in st.session_state:
        st.session_state.start_time = time.time()
    timer_placeholder = st.empty()
    elapsed = time.time() - st.session_state.start_time
    timer_placeholder.markdown(f'<div class="timer-box"><div class="timer-label">⏱️ Session Time</div><div class="timer-value">{int(elapsed // 60)}m {int(elapsed % 60)}s</div></div>', unsafe_allow_html=True)

st.markdown('<p class="subtitle">Upload files and process Transaction data with automated tagging and calculations</p>', unsafe_allow_html=True)
st.divider()

col1, col2 = st.columns(2)
with col1:
    st.markdown("### 📥 1. Upload WS Transaction File")
    input_file = st.file_uploader("Transaction file (Excel)", type=['xlsx'], key="txn")
with col2:
    st.markdown("### 👥 2. Upload System Client Master")
    if input_file:
        system_client_file = st.file_uploader("Client Master (Excel)", type=['xlsx'], key="client")
    else:
        st.warning("⚠️ Upload WS Transaction File first")
        system_client_file = None

col3, col4 = st.columns(2)
with col3:
    st.markdown("### 📋 3. Upload System Scheme Master")
    if input_file and system_client_file:
        system_scheme_file = st.file_uploader("Scheme Master (Excel)", type=['xlsx'], key="scheme")
    else:
        st.warning("⚠️ Upload Client Master first")
        system_scheme_file = None
with col4:
    st.markdown("### 📂 4. Upload MASTER Excel File")
    if input_file and system_client_file and system_scheme_file:
        master_file_raw = st.file_uploader("Main Master file (Excel)", type=['xlsx'], key="master")
    else:
        st.warning("⚠️ Upload Scheme Master first")
        master_file_raw = None

if master_file_raw:
    try:
        proc_start = time.time()
        
        with st.spinner("🚀 Processing data... Please wait..."):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            status_text.text("📂 Loading files...")
            system_client = pd.read_excel(system_client_file)
            system_scheme = pd.read_excel(system_scheme_file)
            master_bytes = master_file_raw.getvalue()
            master_file_io = BytesIO(master_bytes)
            progress_bar.progress(5)
            
            status_text.text("👥 Updating Client Master...")
            master_client = pd.read_excel(master_file_io, sheet_name="Client Master")
            target_columns = ['CLIENTID', 'CLIENTNAME', 'CLIENTCODE', 'PANNUMBER', 'GROUPNAME', 'RELMGRNAME', 'BILLGROUP']
            system_client_normalized = {normalize_col(c): c for c in system_client.columns}
            
            system_matched_cols = {}
            for target_col in target_columns:
                if normalize_col(target_col) in system_client_normalized:
                    system_matched_cols[target_col] = system_client_normalized[normalize_col(target_col)]
            
            if system_matched_cols:
                system_client_filtered = system_client[list(system_matched_cols.values())].copy()
                system_client_filtered.columns = list(system_matched_cols.keys())
            else:
                raise Exception("No matching columns in System Client Master!")
            
            system_client_filtered['_clientcode_clean'] = system_client_filtered['CLIENTCODE'].astype(str).str.strip().str.replace(".0", "", regex=False).str.upper()
            master_client['_clientcode_clean'] = master_client['CLIENTCODE'].astype(str).str.strip().str.replace(".0", "", regex=False).str.upper()
            
            master_clientcodes = set(master_client['_clientcode_clean'].unique())
            system_clientcodes = set(system_client_filtered['_clientcode_clean'].unique())
            missing_clientcodes = system_clientcodes - master_clientcodes
            
            if len(missing_clientcodes) > 0:
                missing_records = system_client_filtered[system_client_filtered['_clientcode_clean'].isin(missing_clientcodes)].copy().drop(columns=['_clientcode_clean'])
                master_columns = [col for col in master_client.columns if col != '_clientcode_clean']
                master_col_normalized = {normalize_col(c): c for c in master_columns}
                
                missing_records_mapped = pd.DataFrame()
                for std_col in missing_records.columns:
                    if normalize_col(std_col) in master_col_normalized:
                        master_col_name = master_col_normalized[normalize_col(std_col)]
                        missing_records_mapped[master_col_name] = missing_records[std_col]
                
                for col in master_columns:
                    if col not in missing_records_mapped.columns:
                        missing_records_mapped[col] = ""
                
                missing_records_mapped = missing_records_mapped[master_columns]
                master_client_updated = pd.concat([master_client.drop(columns=['_clientcode_clean']), missing_records_mapped], ignore_index=True)
            else:
                master_client_updated = master_client.drop(columns=['_clientcode_clean'])
            
            progress_bar.progress(15)
            
            status_text.text("📊 Updating Scheme Master...")
            master_scheme = pd.read_excel(master_file_io, sheet_name="Scheme Master", header=1)
            system_scheme_df = pd.read_excel(system_scheme_file, header=0)
            
            scheme_column_mapping = {'SYMBOLID': 'SYMBOLID', 'SYMBOLNAME': 'Scheme name', 'ISINCODE': 'ISIN', 'REFSYMBOL5': 'Symbolcode5', 'DIMNAME15': 'DIMNAME15 Old', 'ASTCLSNAME': 'ASTCLSNAME', 'DIMNAME13': 'DIMNAME13'}
            system_scheme_normalized = {normalize_col(c): c for c in system_scheme_df.columns}
            system_scheme_matched_cols = {}
            
            for system_col, master_col in scheme_column_mapping.items():
                if normalize_col(system_col) in system_scheme_normalized:
                    system_scheme_matched_cols[master_col] = system_scheme_normalized[normalize_col(system_col)]
            
            if system_scheme_matched_cols:
                system_scheme_filtered = system_scheme_df[list(system_scheme_matched_cols.values())].copy()
                system_scheme_filtered.columns = list(system_scheme_matched_cols.keys())
            else:
                raise Exception("No matching columns in System Scheme Master!")
            
            system_scheme_filtered['_symbolid_clean'] = system_scheme_filtered['SYMBOLID'].astype(str).str.strip().str.upper()
            master_scheme['_symbolid_clean'] = master_scheme['SYMBOLID'].astype(str).str.strip().str.upper()
            
            master_symbolids = set(master_scheme['_symbolid_clean'].unique())
            system_symbolids = set(system_scheme_filtered['_symbolid_clean'].unique())
            missing_symbolids = system_symbolids - master_symbolids
            
            if len(missing_symbolids) > 0:
                missing_scheme_records = system_scheme_filtered[system_scheme_filtered['_symbolid_clean'].isin(missing_symbolids)].copy().drop(columns=['_symbolid_clean'])
                missing_scheme_records['DIMNAME15 - New'] = ""
                missing_scheme_records['ASTCLSNAME New'] = ""
                missing_scheme_records['Manufacturer Name'] = ""
                
                master_scheme_columns = [col for col in master_scheme.columns if col != '_symbolid_clean']
                for col in master_scheme_columns:
                    if col not in missing_scheme_records.columns:
                        missing_scheme_records[col] = ""
                
                missing_scheme_records = missing_scheme_records[master_scheme_columns]
                master_scheme_updated = pd.concat([master_scheme.drop(columns=['_symbolid_clean']), missing_scheme_records], ignore_index=True)
            else:
                master_scheme_updated = master_scheme.drop(columns=['_symbolid_clean'])
            
            progress_bar.progress(25)
            
            status_text.text("💾 Saving Updated Master File...")
            master_file_io.seek(0)
            wb = load_workbook(master_file_io)
            
            if "Client Master" in wb.sheetnames:
                del wb["Client Master"]
            ws_client = wb.create_sheet("Client Master", 0)
            for r in dataframe_to_rows(master_client_updated, index=False, header=True):
                ws_client.append(r)
            
            if "Scheme Master" in wb.sheetnames:
                del wb["Scheme Master"]
            ws_scheme = wb.create_sheet("Scheme Master", 1)
            
            master_file_io.seek(0)
            original_scheme_headers = pd.read_excel(master_file_io, sheet_name="Scheme Master", nrows=1, header=None)
            ws_scheme.append(original_scheme_headers.iloc[0].tolist())
            
            for r in dataframe_to_rows(master_scheme_updated, index=False, header=True):
                ws_scheme.append(r)
            
            master_output = BytesIO()
            wb.save(master_output)
            master_output.seek(0)
            master_file_path = BytesIO(master_output.getvalue())
            progress_bar.progress(35)
            
            status_text.text("📋 Loading Transaction Data...")
            df = pd.read_excel(input_file)
            base_rows = len(df)
            original_cols = df.columns.tolist()
            df = strip_time_from_dates(df)
            
            client_col = find_col(df, ["client name"])
            ws_col = find_col(df, ["ws account code"])
            sec_col = find_col(df, ["security code"])
            trf_col = find_col(df, ["trfamt", "transfer amount"])
            net_col = find_col(df, ["net amount", "amount"])
            txn_col = find_col(df, ["tran desc", "transaction description"])
            desc_col = find_col(df, ["descmemo", "desc memo", "description memo"])
            
            df["Length"] = df[ws_col].astype(str).str.len()
            df["Del Tag"] = ""
            df["Ambit First"] = ""
            progress_bar.progress(40)
            
            status_text.text("🏷️ Applying Ambit First Tags...")
            ambit_first = pd.read_excel(master_file_path, sheet_name="Ambit First")
            ambit_first.columns = ambit_first.columns.astype(str).str.strip().str.lower().str.replace(" ", "_")
            
            df["_ws_clean"] = df[ws_col].astype(str).str.strip().str.replace(".0", "", regex=False)
            ambit_first["_client_clean"] = ambit_first["clientcode"].astype(str).str.strip().str.replace(".0", "", regex=False)
            ambit_first = ambit_first.drop_duplicates(subset=["_client_clean"])
            
            ambit_set = set(ambit_first["_client_clean"])
            matches = df["_ws_clean"].isin(ambit_set)
            df.loc[matches, "Ambit First"] = "Ambit First"
            df["_ws"] = df["_ws_clean"]
            progress_bar.progress(45)
            
            status_text.text("🏷️ Applying Del Tags...")
            pan_condition = (df["Del Tag"] == "") & (df["Length"] == 10)
            df.loc[pan_condition, "Del Tag"] = "Del PAN"
            
            pms_condition = ((df["Del Tag"] == "") & (df["Ambit First"] == "") & df["_ws"].str.upper().str.startswith(("ND", "DS", "DM")))
            df.loc[pms_condition, "Del Tag"] = "Del PMS"
            
            awpl_condition = ((df["Del Tag"] == "") & df[client_col].str.contains("ambit wealth", case=False, na=False))
            df.loc[awpl_condition, "Del Tag"] = "Del AWPL"
            
            afpl_condition = ((df["Del Tag"] == "") & df[client_col].str.contains("Ambit Finvest Private Limited", case=False, na=False))
            df.loc[afpl_condition, "Del Tag"] = "Del AFPL"
            
            dummy_condition = ((df["Del Tag"] == "") & df[client_col].str.contains("dummy", case=False, na=False))
            df.loc[dummy_condition, "Del Tag"] = "Del Dummy"
            
            cash_condition = ((df["Del Tag"] == "") & df[sec_col].str.contains("cash", case=False, na=False))
            df.loc[cash_condition, "Del Tag"] = "Del Cash"
            
            tds_condition = ((df["Del Tag"] == "") & df[sec_col].str.contains("tds", case=False, na=False))
            df.loc[tds_condition, "Del Tag"] = "Del TDSAccount"
            
            mfapp_condition = ((df["Del Tag"] == "") & df[sec_col].str.contains("mfapplication", case=False, na=False))
            df.loc[mfapp_condition, "Del Tag"] = "Del MFApplication"
            
            intaccpur_condition = ((df["Del Tag"] == "") & df[sec_col].str.contains("intaccpur", case=False, na=False))
            df.loc[intaccpur_condition, "Del Tag"] = "Del INTACCPUR"
            progress_bar.progress(55)
            
            status_text.text("📝 Processing Transaction Types...")
            tt = pd.read_excel(master_file_path, sheet_name="Trnx Type Update")
            tt.columns = [str(c).strip() for c in tt.columns]
            
            tran_desc_col = tt.columns[0]
            replace_col = tt.columns[1] if len(tt.columns) > 1 else None
            delete_col = tt.columns[4] if len(tt.columns) > 4 else None
            tt = tt.drop_duplicates(subset=[tran_desc_col])
            
            replace_map = {}
            if replace_col:
                replace_map = dict(zip(tt[tran_desc_col].astype(str).str.strip(), tt[replace_col].astype(str).str.strip()))
            
            delete_lookup = set()
            if delete_col:
                delete_lookup = set(tt[delete_col].dropna().astype(str).str.strip())
            
            df['_txn_clean'] = df[txn_col].astype(str).str.strip()
            df["Revised Trnx Amount"] = np.where(pd.to_numeric(df[trf_col], errors="coerce") > 1, pd.to_numeric(df[trf_col], errors="coerce"), pd.to_numeric(df[net_col], errors="coerce"))
            
            broker_inflow_condition = ((df[desc_col].astype(str).str.strip() == "Broker Change") & (df[txn_col].astype(str).str.strip() == "InFlow"))
            df["Consider"] = df['_txn_clean'].map(replace_map).fillna("")
            df.loc[broker_inflow_condition, "Consider"] = "AUM Trf In"
            
            df["Delete"] = df['_txn_clean'].apply(lambda x: x if x in delete_lookup else "")
            df["Trans Type 2"] = df["Consider"]
            df["Gross Sales"] = np.where(df["Trans Type 2"].isin(["Purchase", "AUM Trf In", "Switch In", "SIP"]), "Gross Sales", "Redemption")
            df["Net Sales"] = np.where(df["Trans Type 2"].isin(["Purchase", "AUM Trf In", "Switch In", "SIP", "Redemption", "AUM Trf Out", "Switch Out", "SWP"]), "Net Sales", "0")
            
            def calculate_amt_in_crs(row):
                revised_amt = row["Revised Trnx Amount"]
                if pd.isna(revised_amt):
                    return 0
                amt_in_crs = revised_amt / 1e7
                if str(row["Gross Sales"]).strip() == "Redemption":
                    amt_in_crs = -amt_in_crs
                return amt_in_crs
            
            df["Amt in Crs"] = df.apply(calculate_amt_in_crs, axis=1)
            progress_bar.progress(65)
            
            status_text.text("🔍 Looking up Scheme Master...")
            scheme = pd.read_excel(master_file_path, sheet_name="Scheme Master", header=1)
            scheme.columns = scheme.columns.str.lower().str.strip()
            scheme = scheme.drop_duplicates(subset=["symbolid"])
            
            df = df.merge(scheme[["symbolid", "dimname15 - new", "astclsname new", "dimname13", "manufacturer name"]], how="left", left_on=sec_col, right_on="symbolid")
            df.rename(columns={"dimname15 - new": "Product New", "astclsname new": "Asset Class New", "dimname13": "Product Category New", "manufacturer name": "Manufacturer Name New"}, inplace=True)
            df.drop(columns=["symbolid"], inplace=True)
            
            ambit_first_mask = (df["Ambit First"] == "Ambit First")
            if ambit_first_mask.sum() > 0:
                df.loc[ambit_first_mask, "Product New"] = "GPC - PMS"
                df.loc[ambit_first_mask, "Asset Class New"] = "Other NDPMS"
                df.loc[ambit_first_mask, "Product Category New"] = "Equity PMS"
                df.loc[ambit_first_mask, "Manufacturer Name New"] = "GPC - Ambit First"
            progress_bar.progress(75)
            
            status_text.text("🔍 Looking up Client Master...")
            client = pd.read_excel(master_file_path, sheet_name="Client Master")
            client.columns = client.columns.str.lower().str.strip().str.replace(" ", "_")
            client["_client_clean"] = client["clientcode"].astype(str).str.strip().str.replace(".0", "", regex=False).str.upper()
            client = client.drop_duplicates(subset=["_client_clean"], keep="last")
            
            df["_ws_upper"] = df["_ws"].str.upper()
            df = df.merge(client[["_client_clean", "groupname", "pannumber", "relmgrname"]], how="left", left_on="_ws_upper", right_on="_client_clean")
            df.rename(columns={"groupname": "Family Name as per Client Master", "pannumber": "Pan No", "relmgrname": "Banker Name"}, inplace=True)
            df.drop(columns=["_client_clean", "_ws_upper"], inplace=True, errors='ignore')
            progress_bar.progress(80)
            
            status_text.text("🔍 Looking up Employee Mapping...")
            emp = pd.read_excel(master_file_path, sheet_name="Employee Mapping Master")
            emp.columns = emp.columns.str.lower().str.strip().str.replace(" ", "_")
            emp = emp.drop_duplicates(subset=["banker_name"])
            
            df = df.merge(emp[["banker_name", "banker_name_new", "banker_group_name", "group_tag"]], how="left", left_on="Banker Name", right_on="banker_name")
            df.rename(columns={"banker_name_new": "Banker Name New", "banker_group_name": "Banker Group Name", "group_tag": "Banker Group Tag"}, inplace=True)
            df.drop(columns=["banker_name"], inplace=True)
            progress_bar.progress(85)
            
            status_text.text("🔍 Looking up NTB Data...")
            ntb = pd.read_excel(master_file_path, sheet_name="NTB Data")
            ntb.columns = ntb.columns.str.lower().str.strip().str.replace(" ", "_")
            ntb = ntb.drop_duplicates(subset=["family_name"])
            
            df = df.merge(ntb[["family_name", "month", "fy"]], how="left", left_on="Family Name as per Client Master", right_on="family_name")
            df.rename(columns={"month": "NTB Month", "fy": "NTB FY"}, inplace=True)
            df.drop(columns=["family_name"], inplace=True)
            
            df["Family Name Final"] = ""
            df["Pk Remark"] = ""
            df["Extra column"] = ""
            df["Month-New"] = ""
            df["YTD Tag"] = ""
            df["Month-New For Banker MIS"] = ""
            progress_bar.progress(90)
            
            status_text.text("📋 Finalizing Data...")
            final_column_order = [*original_cols, "Revised Trnx Amount", "Consider", "Delete", "Trans Type 2", "Gross Sales", "Net Sales", "Product New", "Asset Class New", "Product Category New", "Manufacturer Name New", "Banker Name", "Banker Name New", "Banker Group Name", "Banker Group Tag", "Amt in Crs", "Family Name as per Client Master", "Family Name Final", "Pk Remark", "Extra column", "Month-New", "YTD Tag", "Ambit First", "Pan No", "Month-New For Banker MIS", "NTB Month", "NTB FY", "Length", "Del Tag"]
            final_column_order = [col for col in final_column_order if col in df.columns]
            df = df[final_column_order]
            
            assert len(df) == base_rows, f"Row mismatch! Input={base_rows}, Output={len(df)}"
            
            df_working = df[df["Del Tag"] == ""].copy()
            df_final = df_working[(df_working["Consider"].notna()) & (df_working["Consider"] != "") & (df_working["Delete"].isna() | (df_working["Delete"] == ""))].copy()
            progress_bar.progress(95)
            
            status_text.text("💾 Saving Output Files...")
            mis_output = BytesIO()
            with pd.ExcelWriter(mis_output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Raw Dump', index=False)
                df_working.to_excel(writer, sheet_name='Working', index=False)
                df_final.to_excel(writer, sheet_name='Final', index=False)
            mis_output.seek(0)
            
            progress_bar.progress(100)
            status_text.text("✅ Processing Complete!")
            time.sleep(0.5)
        
        progress_bar.empty()
        status_text.empty()
        processing_time = time.time() - proc_start
        
        st.divider()
        st.subheader("🏁 Processing Results")
        
        m1, m2 = st.columns(2)
        m1.metric("🆕 New Clients Added", len(missing_clientcodes))
        m2.metric("🆕 New Schemes Added", len(missing_symbolids))
        
        r1, r2, r3 = st.columns(3)
        r1.metric("📊 Raw Dump Rows", len(df))
        r2.metric("🛠️ Working Rows", len(df_working))
        r3.metric("🎯 Final Rows", len(df_final))
        
        st.divider()
        
        d1, d2 = st.columns(2)
        d1.download_button("📥 Download Transaction MIS", data=mis_output.getvalue(), file_name="Transaction_MIS_Final.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        d2.download_button("📥 Download Updated Master", data=master_output.getvalue(), file_name="Updated_Master_File.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        
        st.success(f"✅ Processing completed in {processing_time:.2f} seconds!")
        st.balloons()
        
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        st.exception(e)
elif not input_file:
    st.info("👋 Welcome! Start by uploading the **WS Transaction File** in block 1.")
